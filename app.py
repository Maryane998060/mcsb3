"""
IRPF Renda Variável B3 — versão Flask com SQLAlchemy
"""

import os
import uuid
import json
from datetime import date, datetime
from io import BytesIO
from decimal import Decimal
from flask import (
    Flask, render_template, request, redirect,
    url_for, session, send_file, send_from_directory, jsonify,
    flash,
)
from flask_login import (
    LoginManager, login_user, logout_user, login_required,
    current_user,
)

from parsers.excel_parser import parse_negociacao_sheet, parse_movimentacao_sheet
from parsers.pdf_parser import parse_informe_pdf
from parsers.cmpa_calculator import calculate_positions
from parsers.income_calculator import calculate_income, aggregate_by_ir_type, detect_year_from_movements
from parsers.cost_validator import validate_costs
from parsers.cmpa_advanced import CalculadoraCMPA
from parsers.tax_calculator import CalculadoraIR
from parsers.nota_corretagem_parser import parse_notas_corretagem, merge_taxas_nas_posicoes
from parsers.darf_calculator import calcular_darf_mensal, resumo_darf_ano
import auth as auth_module

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'mcs-b3-secret-key-change-in-production-2025')
# ── Flask-Login ────────────────────────────────────────────────────────────────
login_manager = LoginManager(app)
login_manager.login_view  = 'login'
login_manager.login_message = 'Por favor, faça login para acessar o sistema.'
login_manager.login_message_category = 'warning'

@login_manager.user_loader
def load_user(user_id):
    return auth_module.buscar_por_id(user_id)

# Inicializa admin padrão na primeira execução
auth_module.inicializar_admin_se_necessario()

# Cria o usuário SUP padrão se não existir (para deploy em nuvem)
def _garantir_sup_inicial():
    try:
        if not auth_module.buscar_por_username('sup'):
            auth_module.criar_usuario('sup', 'Administrador SUP', 'Ctbc.123', role='sup')
            print("[AUTH] Usuario SUP criado automaticamente.")
    except Exception as e:
        print(f"[AUTH] Aviso ao criar SUP: {e}")

_garantir_sup_inicial()

# Armazenamento em memória dos relatórios processados (chaveado por UUID)
_report_store: dict = {}

ALLOWED_EXCEL = {'xlsx', 'xls'}
ALLOWED_PDF   = {'pdf'}

# ── Carrega ajustes do arquivo de configuração ─────────────────────────────────
_AJUSTES_FILE = os.path.join(os.path.dirname(__file__), 'ajustes_irpf.json')

def _load_ajustes() -> dict:
    """Carrega os ajustes do arquivo ajustes_irpf.json automaticamente."""
    try:
        with open(_AJUSTES_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data.get('ajustes', {})
    except FileNotFoundError:
        return {}
    except Exception as e:
        print(f"[AVISO] Erro ao carregar ajustes_irpf.json: {e}")
        return {}


# ── Rotas de Autenticação ──────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('upload'))

    error    = None
    next_url = request.args.get('next', '')

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        next_url = request.form.get('next', '')

        user = auth_module.buscar_por_username(username)
        if user and user.ativo and user.verificar_senha(password):
            login_user(user, remember=False)
            return redirect(next_url or url_for('upload'))
        else:
            error = 'Usuário ou senha incorretos.'

    return render_template('login.html', error=error, next_url=next_url)


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


# ── Rotas de Gestão de Usuários (somente SUP) ──────────────────────────────────

@app.route('/usuarios')
@login_required
def usuarios():
    if not current_user.is_sup:
        flash('Acesso restrito a superusuários.', 'danger')
        return redirect(url_for('upload'))
    lista = auth_module.listar_usuarios()
    return render_template('usuarios.html', usuarios=lista)


@app.route('/usuarios/criar', methods=['POST'])
@login_required
def usuarios_criar():
    if not current_user.is_sup:
        flash('Acesso restrito a superusuários.', 'danger')
        return redirect(url_for('upload'))
    try:
        username = request.form.get('username', '').strip()
        nome     = request.form.get('nome', '').strip()
        senha    = request.form.get('senha', '')
        role     = request.form.get('role', 'user')
        auth_module.criar_usuario(username, nome, senha, role)
        flash(f'Usuário "{username}" criado com sucesso.', 'success')
    except ValueError as e:
        flash(str(e), 'danger')
    return redirect(url_for('usuarios'))


@app.route('/usuarios/<user_id>/senha', methods=['POST'])
@login_required
def usuarios_senha(user_id):
    if not current_user.is_sup and current_user.id != user_id:
        flash('Sem permissão.', 'danger')
        return redirect(url_for('upload'))
    try:
        nova = request.form.get('nova_senha', '')
        auth_module.alterar_senha(user_id, nova)
        flash('Senha alterada com sucesso.', 'success')
    except ValueError as e:
        flash(str(e), 'danger')
    dest = url_for('usuarios') if current_user.is_sup else url_for('upload')
    return redirect(dest)


@app.route('/usuarios/<user_id>/toggle', methods=['POST'])
@login_required
def usuarios_toggle(user_id):
    if not current_user.is_sup:
        flash('Acesso restrito a superusuários.', 'danger')
        return redirect(url_for('upload'))
    try:
        ativo = auth_module.toggle_ativo(user_id)
        flash(f'Usuário {"ativado" if ativo else "desativado"} com sucesso.', 'success')
    except ValueError as e:
        flash(str(e), 'danger')
    return redirect(url_for('usuarios'))


@app.route('/usuarios/<user_id>/remover', methods=['POST'])
@login_required
def usuarios_remover(user_id):
    if not current_user.is_sup:
        flash('Acesso restrito a superusuários.', 'danger')
        return redirect(url_for('upload'))
    try:
        auth_module.remover_usuario(user_id)
        flash('Usuário removido com sucesso.', 'success')
    except ValueError as e:
        flash(str(e), 'danger')
    return redirect(url_for('usuarios'))


def _allowed(filename: str, extensions: set) -> bool:
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in extensions


# ── Rotas ──────────────────────────────────────────────────────────────────────

@app.route('/templates/css/<path:filename>')
def serve_template_css(filename):
    css_dir = os.path.join(os.path.dirname(__file__), 'templates', 'css')
    return send_from_directory(css_dir, filename, mimetype='text/css')


@app.route('/extract-client-info', methods=['POST'])
@login_required
def extract_client_info():
    """Extrai nome e CPF do PDF da B3 para preencher o formulário automaticamente."""
    informe_file = request.files.get('informe')
    if not informe_file or not informe_file.filename:
        return jsonify({'nome': '', 'cpf': ''})
    try:
        pdf_bytes = BytesIO(informe_file.read())
        info = parse_informe_pdf(pdf_bytes)
        return jsonify({
            'nome': info.get('nome_titular', ''),
            'cpf':  info.get('cpf_titular', ''),
        })
    except Exception:
        return jsonify({'nome': '', 'cpf': ''})


@app.route('/', methods=['GET'])
@login_required
def upload():
    current_year = date.today().year
    return render_template('upload.html', current_year=current_year)
    current_year = date.today().year
    return render_template('upload.html', current_year=current_year)


@app.route('/process', methods=['POST'])
@login_required
def process():
    client_name = request.form.get('client_name', '').strip() or 'Contribuinte'
    cpf         = request.form.get('cpf', '').strip()         or '000.000.000-00'
    year_str    = request.form.get('year', '').strip()
    try:
        year = int(year_str)
    except ValueError:
        year = date.today().year - 1

    negociacao_file   = request.files.get('negociacao')
    movimentacao_file = request.files.get('movimentacao')
    informe_file      = request.files.get('informe')
    notas_file        = request.files.get('notas_corretagem')

    if not movimentacao_file or not movimentacao_file.filename:
        return render_template('upload.html', error='O arquivo de Movimentação é obrigatório.')

    transactions  = []
    movements     = []
    pdf_summary   = None
    notas_summary = None

    try:
        # Movimentação (obrigatório)
        mov_bytes = BytesIO(movimentacao_file.read())
        movements = parse_movimentacao_sheet(mov_bytes)

        # Negociação (opcional)
        if negociacao_file and negociacao_file.filename and _allowed(negociacao_file.filename, ALLOWED_EXCEL):
            neg_bytes = BytesIO(negociacao_file.read())
            transactions = parse_negociacao_sheet(neg_bytes)

        # Notas de corretagem PDF (opcional)
        if notas_file and notas_file.filename and _allowed(notas_file.filename, ALLOWED_PDF):
            notas_bytes = BytesIO(notas_file.read())
            notas_summary = parse_notas_corretagem(notas_bytes, transactions=transactions)

        # Informe de Rendimentos PDF (opcional)
        if informe_file and informe_file.filename and _allowed(informe_file.filename, ALLOWED_PDF):
            pdf_bytes = BytesIO(informe_file.read())
            pdf_summary = parse_informe_pdf(pdf_bytes)
            # Preenche nome/CPF automaticamente se não foram informados pelo usuário
            if pdf_summary.get('nome_titular') and client_name == 'Contribuinte':
                client_name = pdf_summary['nome_titular']
            if pdf_summary.get('cpf_titular') and cpf == '000.000.000-00':
                cpf = pdf_summary['cpf_titular']

    except Exception as e:
        return render_template('upload.html', error=f'Erro ao processar arquivos: {e}')

    # Calcula posições e CMPA
    # Carrega ajustes automaticamente do arquivo ajustes_irpf.json
    ajustes_irpf = _load_ajustes()

    result = calculate_positions(transactions, movements, year=year, ajustes_irpf=ajustes_irpf)
    assets           = result['assets']
    sales            = result['sales']
    corporate_events = result['corporate_events']
    operations       = result['operations']
    audit_log        = result['audit_log']
    validation_issues = result['validation_issues']

    # Aplica taxas das notas de corretagem (opcional — só se o PDF foi enviado)
    if notas_summary and notas_summary.get('resumo_por_ticker'):
        assets = merge_taxas_nas_posicoes(assets, notas_summary['resumo_por_ticker'])
        total_taxas_notas = notas_summary['total_taxas']
        notas_processadas = len(notas_summary['notas'])
        if total_taxas_notas > 0:
            validation_issues.append({
                'level':   'info',
                'message': (
                    f"Notas de corretagem: {notas_processadas} nota(s) processada(s). "
                    f"Total de taxas (emolumentos + liquidação) incorporadas ao custo: "
                    f"R${total_taxas_notas:.2f}. "
                    f"O custo médio ajustado aparece na coluna 'CMPA Ajustado'."
                ),
            })
        if notas_summary.get('erros'):
            erros_notas = [e for e in notas_summary['erros'] if 'sem negocios' not in e]
            for err in erros_notas[:3]:
                validation_issues.append({'level': 'warning', 'message': f"Nota de corretagem: {err}"})
    else:
        notas_summary = None

    # Valida custos — só alerta se cobertura for muito baixa (< 80%)
    cost_validation = validate_costs(assets)
    if cost_validation['action_needed'] and cost_validation.get('coverage_percent', 100) < 80:
        validation_issues.append({
            'level': 'warning',
            'message': (
                f"{cost_validation['zero_cost']} ativo(s) com custo zero detectados. "
                f"Cobertura de custos: {cost_validation['coverage_percent']}%. "
                f"Informe os valores fiscais das bonificações no arquivo ajustes_irpf.json."
            ),
        })

    # Calcula proventos APENAS do ano selecionado pelo usuário
    income_events = calculate_income(movements, year)
    detected_year = year  # sempre o ano escolhido

    # Total de proventos do ano inteiro (inclui ativos vendidos durante o ano)
    # Isso é o que aparece nos cards de resumo e bate com o PDF da B3
    aggregated_total_ano = aggregate_by_ir_type(income_events)

    # Filtra vendas apenas do ano selecionado (negociações históricas não são vendas de 2025)
    def _year_of(date_str: str) -> int:
        parts = str(date_str).split('/')
        if len(parts) == 3:
            try: return int(parts[2])
            except: pass
        return 0

    sales_year = [s for s in sales if _year_of(s['date']) == year]

    # ── Filtra ativos usando o PDF como fonte de verdade ──────────────────────
    # Se o PDF foi enviado, usa a posição de ações do PDF para filtrar a carteira.
    # Ativos que não aparecem no PDF foram zerados antes do fim do ano (corretoras antigas).
    # Se o PDF não foi enviado, mantém todos os ativos calculados.
    pdf_positions = pdf_summary.get('inferred_assets', {}) if pdf_summary else {}

    if pdf_positions:
        # Normaliza tickers do PDF (ELET6 → AXIA6 via tabela de conversão)
        from parsers.excel_parser import _normalize_ticker
        pdf_tickers_norm = {_normalize_ticker(t): qty for t, qty in pdf_positions.items()}

        # Tickers que aparecem nos proventos do PDF com JCP ou Restituição de Capital
        # (indica posse no ano — dividendo sozinho não garante posse em 31/12)
        pdf_jcp_rest_tickers = {
            _normalize_ticker(p['ticker'])
            for p in pdf_summary.get('proventos', [])
            if p.get('ir_type') == 10 or 'restituição' in p.get('type', '').lower() or 'restituicao' in p.get('type', '').lower()
        }

        # Mantém ativos que:
        # 1) Aparecem na posição de ações do PDF, OU
        # 2) Têm qty > 0 calculada E têm JCP/Restituição no PDF
        #    (ex: VIVT3 que tem qty mas pode não aparecer na posição por erro do parser)
        # EXCLUÍDO: ativos com apenas dividendo no PDF mas sem posição
        #    (ex: RAIZ4 emprestado — recebeu dividendo mas saiu da custódia)
        assets_ativos = [
            a for a in assets
            if a['ticker'] in pdf_tickers_norm
            or (a['ticker'] in pdf_jcp_rest_tickers and a['quantity'] > 0)
        ]

        # ── Ativos no PDF mas fora do cálculo (histórico de compra faltando) ──
        # Ex: ROXO34 comprado em outra corretora — aparece no PDF mas não nas planilhas
        calc_tickers = {a['ticker'] for a in assets_ativos}
        for ticker_pdf, qty_pdf in pdf_tickers_norm.items():
            if ticker_pdf not in calc_tickers and qty_pdf > 0:
                # Cria entrada com custo zero e aviso
                from parsers.cmpa_calculator import _detect_type, _get_irpf_code
                asset_type = _detect_type(ticker_pdf)
                irpf = _get_irpf_code(asset_type)
                assets_ativos.append({
                    'ticker':           ticker_pdf,
                    'reference_ticker': ticker_pdf,
                    'full_name':        ticker_pdf,
                    'quantity':         qty_pdf,
                    'average_cost':     0.0,
                    'total_cost':       0.0,
                    'cost_prev_year':   0.0,
                    'qty_prev_year':    0.0,
                    'cost_curr_year':   0.0,
                    'dividends':        0.0,
                    'jcp':              0.0,
                    'fii_income':       0.0,
                    'institution':      '',
                    'asset_cnpj':       None,
                    'broker_cnpj':      None,
                    'type':             asset_type,
                    'irpf_grupo':       irpf['grupo'],
                    'irpf_codigo':      irpf['codigo'],
                    'irpf_code':        irpf['code'],
                    'first_purchase_date': '',
                    'last_purchase_date':  '',
                    'first_purchase_quantity': 0.0,
                    'year_end_total':   0.0,
                    'buy_events':       [],
                    'bonus_events':     [],
                    'corporate_events_history': [],
                    'discriminacao':    f'{qty_pdf} AÇÕES {ticker_pdf} — CUSTO NÃO LOCALIZADO. '
                                        f'O histórico de compra não foi encontrado nas planilhas enviadas. '
                                        f'Informe o custo manualmente no arquivo ajustes_irpf.json.',
                })
                validation_issues.append({
                    'level':   'warning',
                    'message': (
                        f"{ticker_pdf}: ativo confirmado pelo PDF ({qty_pdf} ações) mas sem histórico de compra "
                        f"nas planilhas. Custo declarado como R$ 0,00. "
                        f"Informe o custo real em ajustes_irpf.json usando 'custo_declarado'."
                    ),
                })

        # Reordena por ticker
        assets_ativos.sort(key=lambda a: a['ticker'])

        # Valida quantidade vs PDF para ativos que constam na posição
        for a in assets_ativos:
            pdf_qty = pdf_tickers_norm.get(a['ticker'])
            if pdf_qty and abs(a['quantity'] - pdf_qty) > 0.1:
                validation_issues.append({
                    'level':   'warning',
                    'message': (
                        f"Quantidade de {a['ticker']} calculada ({a['quantity']:.2f}) "
                        f"difere do PDF da B3 ({pdf_qty:.2f}). "
                        f"Pode haver eventos corporativos ou vendas não incluídos nos arquivos."
                    ),
                })
    else:
        assets_ativos = assets

    # ── Distribui proventos do ano nos ativos da carteira filtrada ─────────────
    # income_map: só para os ativos em carteira (para a coluna por ativo na tabela)
    # aggregated_ativos: total para validação cruzada com PDF
    # aggregated_total_ano: total real do ano para os cards de resumo (inclui vendidos)
    asset_tickers = {a['ticker'] for a in assets_ativos}
    income_map: dict = {}
    for ev in income_events:
        if ev['ticker'] not in asset_tickers:
            continue
        bucket = income_map.setdefault(ev['ticker'], {'dividends': 0.0, 'jcp': 0.0, 'fii_income': 0.0})
        if ev['ir_type'] == 9:
            bucket['dividends']  += ev['value']
        elif ev['ir_type'] == 10:
            bucket['jcp']        += ev['value']
        elif ev['ir_type'] == 26:
            bucket['fii_income'] += ev['value']

    # Totaliza proventos dos ativos em carteira (para validação cruzada com PDF)
    aggregated_ativos = aggregate_by_ir_type([
        ev for ev in income_events if ev['ticker'] in asset_tickers
    ])

    # Aplica proventos e gera discriminação para cada ativo
    for asset in assets_ativos:
        bucket = income_map.get(asset['ticker'])
        if bucket:
            asset['dividends']  = round(bucket['dividends'],  2)
            asset['jcp']        = round(bucket['jcp'],        2)
            asset['fii_income'] = round(bucket['fii_income'], 2)
        asset['discriminacao'] = _build_discriminacao(asset)

    # ── Validações cruzadas com PDF (feita aqui pois agora aggregated_ativos existe) ──
    if pdf_summary:
        # Usa o total do ano inteiro para validar contra o PDF (que inclui ativos vendidos)
        tax9  = aggregated_total_ano['tax9']
        tax10 = aggregated_total_ano['tax10']
        tax26 = aggregated_total_ano['tax26']
        pdf_tax9  = pdf_summary.get('dividendos') or 0
        pdf_tax10 = pdf_summary.get('jcp') or 0
        pdf_tax26 = pdf_summary.get('rendimento_fii') or 0
        if pdf_tax9  and abs(pdf_tax9  - tax9)  > 0.5:
            validation_issues.append({'level': 'warning',
                'message': f"Dividendos PDF (R${pdf_tax9:.2f}) ≠ calculado (R${tax9:.2f})."})
        if pdf_tax10 and abs(pdf_tax10 - tax10) > 0.5:
            validation_issues.append({'level': 'warning',
                'message': f"JCP PDF (R${pdf_tax10:.2f}) ≠ calculado (R${tax10:.2f})."})
        if pdf_tax26 and abs(pdf_tax26 - tax26) > 0.5:
            validation_issues.append({'level': 'warning',
                'message': f"Rendimento FII PDF (R${pdf_tax26:.2f}) ≠ calculado (R${tax26:.2f})."})
        for note in pdf_summary.get('notes', []):
            validation_issues.append({'level': 'warning', 'message': note})

    # Ativos zerados = estavam no cálculo mas não no PDF (ou foram vendidos no ano)
    tickers_em_assets = {a['ticker'] for a in assets_ativos}
    tickers_zerados_no_ano = []
    for s in sales_year:
        if s['ticker'] not in tickers_em_assets:
            if not any(t['ticker'] == s['ticker'] for t in tickers_zerados_no_ano):
                tickers_zerados_no_ano.append({
                    'ticker': s['ticker'],
                    'motivo': 'Vendido em ' + str(year)
                })

    total_assets_cost    = sum(a['total_cost'] for a in assets_ativos)
    total_prev_year_cost = sum(a.get('cost_prev_year', 0.0) for a in assets_ativos)
    total_gain  = sum(s['gain'] for s in sales_year if s['gain'] > 0)
    total_loss  = sum(s['gain'] for s in sales_year if s['gain'] < 0)
    count_gain  = sum(1 for s in sales_year if s['gain'] > 0)
    count_loss  = sum(1 for s in sales_year if s['gain'] < 0)

    # ── Calcula DARF mensal ────────────────────────────────────────────────────
    # Monta mapa ticker → tipo para diferenciação FII/ETF/BDR/Ação
    asset_type_map = {a['ticker']: a['type'] for a in assets_ativos}
    # Inclui também ativos que foram vendidos (podem não estar em assets_ativos)
    for a in assets:
        if a['ticker'] not in asset_type_map:
            asset_type_map[a['ticker']] = a.get('type', 'Ação')

    darfs_mensais = calcular_darf_mensal(sales_year, asset_map=asset_type_map)
    darf_resumo   = resumo_darf_ano(darfs_mensais)

    # Agrupa vendas do ano selecionado por mês para o relatório DARF
    monthly_sales: dict = {}
    for s in sales_year:
        parts = s['date'].split('/')
        if len(parts) == 3:
            key = f"{parts[1]}/{parts[2]}"
            bucket = monthly_sales.setdefault(key, {'gain': 0.0, 'loss': 0.0, 'net': 0.0, 'count': 0})
            if s['gain'] >= 0:
                bucket['gain'] += s['gain']
            else:
                bucket['loss'] += s['gain']
            bucket['net']   += s['gain']
            bucket['count'] += 1

    # Ranking de proventos por ativo (para o gráfico/ranking na aba Proventos)
    income_ranking: list = []
    _rank_map: dict = {}
    for ev in income_events:
        if ev['ticker'] in asset_tickers:
            _rank_map[ev['ticker']] = _rank_map.get(ev['ticker'], 0.0) + ev['value']
    if _rank_map:
        max_rank = max(_rank_map.values())
        income_ranking = [
            {'ticker': t, 'value': round(v, 2), 'pct': round(v / max_rank * 100)}
            for t, v in sorted(_rank_map.items(), key=lambda x: x[1], reverse=True)
        ]

    report = {
        'client_name':          client_name,
        'cpf':                  cpf,
        'year':                 year,
        'detected_year':        detected_year,
        'assets':               assets_ativos,
        'assets_zerados':       tickers_zerados_no_ano,
        'income_events':        income_events,  # todos do ano (inclui ativos vendidos)
        'income_ranking':       income_ranking,
        'sales':                sales_year,
        'monthly_sales':        monthly_sales,
        'corporate_events':     corporate_events,
        'operations':           operations,
        # Audit log: só mostra eventos relevantes (exclui transferências que são ruído)
        'audit_log':            [
            a for a in audit_log
            if a.get('event', '') not in ('Transferência', 'Custódia', 'Liquidação')
            and 'Transferência/Custódia' not in a.get('reason', '')
        ],
        # Validation issues: filtra avisos de posição insuficiente em anos anteriores ao selecionado
        # (ocorrem quando o histórico começa em 2020 mas o investidor tinha posições em 2019)
        'validation_issues':    [
            v for v in validation_issues
            if not (
                'posição tem apenas' in v.get('message', '')
                and v.get('date', '9999')[-4:] < str(year)
            )
        ],
        'pdf_summary':          pdf_summary,
        'notas_corretagem':     notas_summary,
        'darfs':                darfs_mensais,
        'darf_resumo':          darf_resumo,
        'summary': {
            'total_assets_cost':      round(total_assets_cost, 2),
            'total_prev_year_cost':   round(total_prev_year_cost, 2),
            # Proventos do ano INTEIRO (inclui ativos vendidos) — para bater com PDF
            'total_income':           aggregated_total_ano['total'],
            'tax9':                   aggregated_total_ano['tax9'],
            'tax10':                  aggregated_total_ano['tax10'],
            'tax26':                  aggregated_total_ano['tax26'],
            # Subtotais por ativo em carteira (para a tabela e validação cruzada)
            'tax9_carteira':          aggregated_ativos['tax9'],
            'tax10_carteira':         aggregated_ativos['tax10'],
            'tax26_carteira':         aggregated_ativos['tax26'],
            'total_gain':             round(total_gain, 2),
            'total_loss':             round(total_loss, 2),
            'net_gain':               round(total_gain + total_loss, 2),
            'count_gain':             count_gain,
            'count_loss':             count_loss,
            'sales_year':             year,
        },
        'diagnostics': {
            'negociacao_rows':    len(transactions),
            'movimentacao_rows':  len(movements),
            'assets_found':       len(assets),
            'income_found':       len(income_events),
            'sales_year_count':   len(sales_year),
            'sales_total_count':  len(sales),
        },
    }

    report_id = str(uuid.uuid4())
    _report_store[report_id] = report
    session['report_id'] = report_id
    return redirect(url_for('report'))


@app.route('/report')
@login_required
def report():
    report_id = session.get('report_id')
    if not report_id or report_id not in _report_store:
        return redirect(url_for('upload'))
    return render_template('report.html', r=_report_store[report_id])


@app.route('/new-report')
@login_required
def new_report():
    session.pop('report_id', None)
    return redirect(url_for('upload'))


@app.route('/export/excel')
@login_required
def export_excel():
    report_id = session.get('report_id')
    if not report_id or report_id not in _report_store:
        return redirect(url_for('upload'))

    r = _report_store[report_id]
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    header_font  = Font(bold=True, color='FFFFFF')
    header_fill  = PatternFill('solid', fgColor='1e3a5f')
    center_align = Alignment(horizontal='center')

    def write_sheet(ws, title, headers, rows):
        ws.title = title
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font   = header_font
            cell.fill   = header_fill
            cell.alignment = center_align
        for row_idx, row in enumerate(rows, 2):
            for col_idx, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    # Aba 1 — Bens & Direitos
    ws1 = wb.active
    write_sheet(ws1, 'Bens e Direitos', [
        'Ticker', 'Nome', 'Tipo', 'Código IRPF', 'Quantidade',
        'CMPA (R$)', 'Custo Total (R$)', 'Dividendos (R$)', 'JCP (R$)',
        'Rendimento FII (R$)', 'CNPJ Ativo', 'Corretora',
    ], [
        [
            a['ticker'], a['full_name'], a['type'], a['irpf_code'],
            a['quantity'], a['average_cost'], a['total_cost'],
            a['dividends'], a['jcp'], a['fii_income'],
            a.get('asset_cnpj', ''), a['institution'],
        ]
        for a in r['assets']
    ])

    # Aba 2 — Proventos
    ws2 = wb.create_sheet()
    write_sheet(ws2, 'Proventos', [
        'Data', 'Ticker', 'Tipo', 'Tipo IR', 'Quantidade', 'Preço Unit.', 'Valor (R$)',
    ], [
        [e['date'], e['ticker'], e['type'], e['ir_type'], e['quantity'], e['unit_price'], e['value']]
        for e in r['income_events']
    ])

    # Aba 3 — Vendas
    ws3 = wb.create_sheet()
    write_sheet(ws3, 'Lucro e Prejuízo', [
        'Data', 'Ticker', 'Quantidade', 'Preço Venda (R$)',
        'CMPA (R$)', 'Base de Custo (R$)', 'Ganho/Perda (R$)',
    ], [
        [s['date'], s['ticker'], s['quantity'], s['sale_price'],
         s['cmpa_at_sale'], s['cost_basis'], s['gain']]
        for s in r['sales']
    ])

    # Aba 4 — Eventos Corporativos
    ws4 = wb.create_sheet()
    write_sheet(ws4, 'Eventos Corporativos', [
        'Data', 'Ticker', 'Tipo', 'Descrição', 'Δ Quantidade',
    ], [
        [e['date'], e['ticker'], e['type'], e['details'], e['quantity_change']]
        for e in r['corporate_events']
    ])

    # Aba 5 — Operações
    ws5 = wb.create_sheet()
    write_sheet(ws5, 'Operações', [
        'Data', 'Ticker', 'Nome', 'Tipo', 'Quantidade',
        'Preço Unit.', 'Valor (R$)', 'Corretora', 'Nota',
    ], [
        [
            o['date'], o['ticker'], o.get('full_name', ''), o['type'],
            o['quantity'], o.get('unit_price', 0), o.get('value', 0),
            o.get('institution', ''), o.get('note', ''),
        ]
        for o in r['operations']
    ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"IRPF_B3_{r['client_name'].replace(' ', '_')}_{r['year']}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


@app.route('/darf')
@login_required
def darf():
    """Relatório DARF (imposto de renda)"""
    report_id = session.get('report_id')
    if not report_id or report_id not in _report_store:
        return jsonify({'error': 'Nenhum relatório encontrado'}), 404
    
    r = _report_store[report_id]
    calc_ir = CalculadoraIR()
    
    for s in r['sales']:
        calc_ir.adicionar_venda(
            data=_parse_date_to_datetime(s['date']),
            ticker=s['ticker'],
            quantidade=Decimal(str(s['quantity'])),
            preco_compra=Decimal(str(s.get('cmpa_at_sale', 0))),
            preco_venda=Decimal(str(s['sale_price'])),
            custo_medio_aquisicao=Decimal(str(s.get('cmpa_at_sale', 0))),
            custos_operacionais=Decimal('0'),
        )
    
    darfs_gerados = calc_ir.gerar_darf_completo(r['year'])
    
    total_imposto = sum(Decimal(str(d['valor_imposto'])) for d in darfs_gerados)
    total_lucro = sum(Decimal(str(d['valor_lucro'])) for d in darfs_gerados)
    
    return jsonify({
        'cliente': {'nome': r['client_name'], 'cpf': r['cpf']},
        'ano_calendario': r['year'],
        'total_lucro': float(total_lucro),
        'total_imposto': float(total_imposto),
        'meses': darfs_gerados
    })


# ─── Helper functions for reports ────────────────────────────────────────────────

def _fmt_qty(n: float) -> str:
    """Formata quantidade no padrão BR (vírgula decimal, sem zeros desnecessários)."""
    if n == int(n):
        return str(int(n))
    s = f"{n:.4f}".rstrip('0')
    return s.replace('.', ',')


def _fmt_brl(n: float) -> str:
    """Formata valor monetário no padrão BR: 1.234,56."""
    return f"{n:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')


def _build_discriminacao(a: dict) -> str:
    """
    Monta a discriminação inteiramente dinâmica e limpa para qualquer cliente,
    baseando-se exclusivamente nos metadados e histórico calculados a partir das planilhas.
    """
    unidade_map = {'FII': 'COTAS', 'ETF': 'COTAS', 'BDR': 'BDRs'}
    unidade  = unidade_map.get(a.get('type'), 'AÇÕES')
    ticker   = a['ticker']
    qty_str  = _fmt_qty(a['quantity'])
    cmpa_br  = _fmt_brl(a['average_cost'])

    # Extração limpa e direta da instituição financeira e CNPJ mapeados na planilha
    inst_name = str(a.get('institution', '')).upper().strip()
    cnpj_corretora = a.get('broker_cnpj', '00.000.000/0000-00')

    custodia = f"CUSTODIADAS NA {inst_name if inst_name else 'CORRETORA DE CUSTÓDIA'}"
    if cnpj_corretora and cnpj_corretora != '00.000.000/0000-00':
        custodia += f", CNPJ {cnpj_corretora}"

    buy_events = a.get('buy_events', [])
    if len(buy_events) == 1:
        datas_str = f"EM {buy_events[0]['date']}"
    elif buy_events:
        datas = list(dict.fromkeys(b['date'] for b in buy_events))
        datas_str = f"ENTRE {datas[0]} E {datas[-1]}" if len(datas) > 1 else f"EM {datas[0]}"
    else:
        first_date = a.get('first_purchase_date', '').strip()
        if first_date:
            datas_str = f"DESDE {first_date}"
        else:
            datas_str = "CONFORME HISTÓRICO DA B3"

    texto = f"{qty_str} {unidade} {ticker}, ADQUIRIDAS {datas_str}, COM O CUSTO MÉDIO DE R$ {cmpa_br}, {custodia}."

    # Processamento condicional de eventos societários extraídos do Ledger real do cliente
    ajustes = []
    
    for bns in a.get('bonus_events', []):
        v_fiscal = bns.get('unit_fiscal_value', 0)
        v_fiscal_str = f" PELO VALOR UNITÁRIO FISCAL DE R$ {_fmt_brl(v_fiscal)}" if v_fiscal > 0 else ""
        ajustes.append(f"RECEBIMENTO DE BONIFICAÇÃO DE {bns.get('quantity_change', bns.get('quantity'))} ATIVOS EM {bns['date']}{v_fiscal_str}")
        
    for evt in a.get('corporate_events_history', []):
        if evt.get('type') == 'GRUPAMENTO':
            ajustes.append(f"POSIÇÃO AJUSTADA POR EVENTO CORPORATIVO DE GRUPAMENTO EM {evt['date']}")
        elif evt.get('type') == 'RESTITUICAO_CAPITAL':
            val_rest = evt.get('value', 0)
            ajustes.append(f"REDUÇÃO DE CUSTO HISTÓRICO DEVIDO A RESTITUIÇÃO DE CAPITAL EM DINHEIRO NO VALOR DE R$ {_fmt_brl(val_rest)} EM {evt['date']}")

    if ajustes:
        texto += " POSIÇÃO AJUSTADA NO ANO BASE DEVIDO A: " + "; ".join(ajustes) + "."

    return texto


def _parse_date_to_datetime(date_str: str):
    """Converte string DD/MM/YYYY ou ISO de forma robusta para date"""
    if not date_str:
        return None
    date_str = str(date_str).strip().split(' ')[0] # Limpa horas residuais se houver
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y'):
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None
    

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)