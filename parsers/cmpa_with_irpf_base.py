"""
Calculador de CMPA com base em IRPF ANTERIOR + Transações 2025

Algoritmo correto:
1. Lê posições iniciais do IRPF anterior (posição em 31/12/2024)
2. Aplica compras/vendas/eventos de 2025
3. Calcula novo CMPA
"""

from datetime import datetime


def extract_irpf_positions(irpf_file: str) -> dict:
    """
    Extrai posições finais do IRPF anterior.
    Retorna: {ticker: {quantidade: float, custo_total: float, custo_medio: float}}
    """
    import openpyxl
    
    positions = {}
    
    try:
        wb = openpyxl.load_workbook(irpf_file, data_only=True)
        ws = wb['Bens e Direitos IRPF']
        
        # Encontra o header
        header_row = None
        for row_idx in range(1, min(10, ws.max_row + 1)):
            cell_val = ws.cell(row_idx, 3).value
            if cell_val and 'B3' in str(cell_val):
                header_row = row_idx
                break
        
        if not header_row:
            print("AVISO: Nao encontrou header no IRPF")
            return positions
        
        # Coluna mapping
        col_ticker = 3      # C: Código B3
        col_qty = 5         # E: Qtd
        col_cost_medio = 6  # F: Custo Médio R$
        
        # Lê dados
        for row_idx in range(header_row + 1, ws.max_row + 1):
            ticker = ws.cell(row_idx, col_ticker).value
            qty = ws.cell(row_idx, col_qty).value
            cost_medio = ws.cell(row_idx, col_cost_medio).value
            
            if not ticker or qty is None:
                continue
            
            ticker = str(ticker).strip().upper()
            qty = float(qty) if qty else 0
            cost_medio = float(cost_medio) if cost_medio else 0
            
            positions[ticker] = {
                'quantidade': qty,
                'custo_medio': cost_medio,
                'custo_total': qty * cost_medio,
            }
        
        wb.close()
    
    except Exception as e:
        print(f"ERRO ao ler IRPF: {e}")
    
    return positions


def apply_transactions_2025(
    irpf_positions: dict,
    transactions: list,
    movements: list,
    pdf_positions: dict = None
) -> dict:
    """
    Aplica transações de 2025 às posições do IRPF anterior.
    Recalcula CMPA.
    """
    
    from collections import defaultdict
    
    # Inicializa com posições do IRPF
    assets = {}
    for ticker, pos in irpf_positions.items():
        assets[ticker] = {
            'ticker': ticker,
            'quantidade_inicial': pos['quantidade'],
            'quantidade_atual': pos['quantidade'],
            'custo_total_inicial': pos['custo_total'],
            'custo_total_atual': pos['custo_total'],
            'custo_medio': pos['custo_medio'],
            'transacoes_2025': [],
        }
    
    # Agrupa transações por ticker
    trans_by_ticker = defaultdict(list)
    for trans in transactions:
        ticker = trans.get('ticker', '').upper()
        date = trans.get('date', '')
        
        # Filtra apenas 2025
        if date and '2025' in date and ticker:
            trans_by_ticker[ticker].append(trans)
    
    # Processa transações de 2025
    for ticker, trans_list in trans_by_ticker.items():
        # Cria asset se nao existir (compra de novo ticker)
        if ticker not in assets:
            assets[ticker] = {
                'ticker': ticker,
                'quantidade_inicial': 0,
                'quantidade_atual': 0,
                'custo_total_inicial': 0,
                'custo_total_atual': 0,
                'custo_medio': 0,
                'transacoes_2025': [],
            }
        
        asset = assets[ticker]
        qtd_atual = asset['quantidade_inicial']
        custo_atual = asset['custo_total_inicial']
        
        # Processa cada transação 2025
        for trans in sorted(trans_list, key=lambda x: x.get('date', '')):
            tipo = trans.get('type', '').lower()
            qty = float(trans.get('quantity', 0))
            preco = float(trans.get('price', 0))
            valor = qty * preco
            
            if 'compra' in tipo or 'compra' in tipo:
                # COMPRA: aumenta quantidade e custo
                novo_custo_medio = (custo_atual + valor) / (qtd_atual + qty) if (qtd_atual + qty) > 0 else 0
                
                asset['transacoes_2025'].append({
                    'data': trans.get('date', ''),
                    'tipo': tipo,
                    'quantidade': qty,
                    'preco': preco,
                    'valor': valor,
                    'acao': 'compra'
                })
                
                qtd_atual += qty
                custo_atual += valor
            
            elif 'venda' in tipo or 'venda' in tipo:
                # VENDA: reduz quantidade, mas custo fica mesmo
                # (usa CMPA anterior para calcular ganho/perda)
                
                asset['transacoes_2025'].append({
                    'data': trans.get('date', ''),
                    'tipo': tipo,
                    'quantidade': qty,
                    'preco': preco,
                    'valor': valor,
                    'acao': 'venda'
                })
                
                qtd_atual -= qty
                # Custo sai proporcional
                if asset['quantidade_inicial'] + sum(t['quantidade'] for t in asset['transacoes_2025'] if 'compra' in t['acao']) > 0:
                    custo_vendido = asset['custo_medio'] * qty
                    custo_atual -= custo_vendido
        
        # Atualiza valores finais
        asset['quantidade_atual'] = qtd_atual
        asset['custo_total_atual'] = custo_atual
        asset['custo_medio'] = custo_atual / qtd_atual if qtd_atual > 0 else 0
        
        # Se PDF disponível, valida quantidade
        if pdf_positions and ticker in pdf_positions:
            pdf_qty = pdf_positions[ticker].get('quantidade', 0)
            if abs(qtd_atual - pdf_qty) > 0.01:
                asset['_warning'] = f"Qty atual ({qtd_atual:.2f}) != PDF ({pdf_qty:.2f})"
    
    return assets


if __name__ == '__main__':
    print("Teste do novo calculador com IRPF anterior...")
    
    # Simula dados
    irpf_pos = {
        'SOJA3': {'quantidade': 200, 'custo_total': 2500, 'custo_medio': 12.50},
        'AGRO3': {'quantidade': 100, 'custo_total': 2499, 'custo_medio': 24.99},
    }
    
    trans_2025 = [
        {
            'ticker': 'SOJA3',
            'type': 'Venda',
            'date': '10/12/2025',
            'quantity': 100,
            'price': 9.08,
        },
    ]
    
    result = apply_transactions_2025(irpf_pos, trans_2025, [])
    
    for ticker, asset in result.items():
        print(f"\n{ticker}:")
        print(f"  Inicial: {asset['quantidade_inicial']:.2f} x R$ {asset['custo_medio']:.2f}")
        print(f"  Final:   {asset['quantidade_atual']:.2f} x R$ {asset['custo_medio']:.2f}")
        if asset['transacoes_2025']:
            print(f"  Transacoes 2025: {len(asset['transacoes_2025'])}")
