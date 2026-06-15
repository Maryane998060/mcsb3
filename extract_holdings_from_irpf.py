"""
Script para extrair dados de Bens & Direitos do IRPF anterior (arquivo de análise)
e gerar um arquivo inicial de posições com custos corretos para inicializar a análise.
"""

import openpyxl
from pathlib import Path
from datetime import datetime, timedelta


def extract_from_irpf_file(file_path: str) -> dict:
    """
    Lê arquivo IRPF anterior e extrai posições com custos.
    Retorna dict mapeando ticker -> {qtd, custo_medio, custo_total}
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb['Bens e Direitos IRPF']
    
    holdings = {}
    
    # Procura o header (row com "Grupo", "Código B3", "Qtd", "Custo Médio")
    header_row = None
    for row_idx in range(1, min(10, ws.max_row + 1)):
        cell_val = ws.cell(row_idx, 3).value
        if cell_val and 'B3' in str(cell_val):
            header_row = row_idx
            break
    
    if not header_row:
        raise ValueError("Não encontrou header com 'Código B3'")
    
    # Coluna mapping (assumindo formato fixo)
    col_ticker = 3      # C: Código B3
    col_qty = 5         # E: Qtd
    col_cost_medio = 6  # F: Custo Médio R$
    
    # Lê dados de ativos
    for row_idx in range(header_row + 1, ws.max_row + 1):
        ticker = ws.cell(row_idx, col_ticker).value
        qty = ws.cell(row_idx, col_qty).value
        cost_medio = ws.cell(row_idx, col_cost_medio).value
        
        if not ticker or not qty or cost_medio is None:
            continue
        
        ticker = str(ticker).strip()
        qty = float(qty) if qty else 0
        cost_medio = float(cost_medio) if cost_medio else 0
        
        holdings[ticker] = {
            'quantity': qty,
            'average_cost': cost_medio,
            'total_cost': qty * cost_medio
        }
    
    wb.close()
    return holdings


def generate_synthetic_negotiations(holdings: dict, target_date: str = '31/12/2024') -> list:
    """
    Gera negociações sintéticas para cada ativo em holdings.
    Cria uma entrada única por ativo com a data fornecida.
    """
    negotiations = []
    
    # Data um dia antes da data alvo (para ser a "última compra")
    parts = target_date.split('/')
    last_buy_date = datetime(int(parts[2]), int(parts[1]), int(parts[0])) - timedelta(days=1)
    buy_date = last_buy_date.strftime('%d/%m/%Y')
    
    for ticker, data in holdings.items():
        qty = data['quantity']
        total = data['total_cost']
        price = data['average_cost']
        
        if qty <= 0 or total <= 0:
            continue
        
        # Cria entrada de negociação sintética
        negotiations.append({
            'date': buy_date,
            'type': 'Compra',
            'market': 'BOVESPA',
            'institution': 'Corretora Histórica',
            'ticker': ticker,
            'full_name': f'{ticker} - Posição Histórica',
            'quantity': qty,
            'price': price,
            'value': total,
            'asset_cnpj': None,
            'broker_cnpj': None,
        })
    
    return negotiations


def save_negotiations_to_excel(negotiations: list, output_path: str):
    """
    Salva negociações em arquivo Excel no formato esperado pela aplicação.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Negociação'
    
    # Headers
    headers = ['Data do Negócio', 'Tipo de Movimentação', 'Mercado', 'Instituição', 
               'Código de Negociação', 'Discriminação', 'Quantidade', 'Preço', 'Valor']
    for col_idx, header in enumerate(headers, 1):
        ws.cell(1, col_idx, header)
    
    # Data
    for row_idx, neg in enumerate(negotiations, 2):
        ws.cell(row_idx, 1, neg['date'])
        ws.cell(row_idx, 2, neg['type'])
        ws.cell(row_idx, 3, neg['market'])
        ws.cell(row_idx, 4, neg['institution'])
        ws.cell(row_idx, 5, neg['ticker'])
        ws.cell(row_idx, 6, neg['full_name'])
        ws.cell(row_idx, 7, neg['quantity'])
        ws.cell(row_idx, 8, neg['price'])
        ws.cell(row_idx, 9, neg['value'])
    
    wb.save(output_path)
    print(f'Arquivo salvo: {output_path}')
    print(f'Total de negociações: {len(negotiations)}')


if __name__ == '__main__':
    input_file = r'g:\Maryane\calculainvestimentoir\exemplos\analise_acoes_IRPF2026_paulo_henrique.xlsx'
    output_file = r'g:\Maryane\calculainvestimentoir\exemplos\negociacao-historico-inicial.xlsx'
    
    print('Extraindo dados do arquivo IRPF...')
    holdings = extract_from_irpf_file(input_file)
    print(f'Encontrados {len(holdings)} ativos:')
    for ticker, data in holdings.items():
        print(f'  {ticker:10} Qty: {data["quantity"]:10.2f} Custo Médio: R$ {data["average_cost"]:10.2f} Total: R$ {data["total_cost"]:12.2f}')
    
    print(f'\nGerando negociações sintéticas...')
    negotiations = generate_synthetic_negotiations(holdings)
    
    print(f'\nSalvando arquivo...')
    save_negotiations_to_excel(negotiations, output_file)
    
    print(f'\nFeito! Use o arquivo {output_file} como base para o cálculo.')
